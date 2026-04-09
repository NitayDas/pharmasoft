import csv
import io
import re
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Prefetch
from django.db.models import Q
from rest_framework import status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Product, PurchaseImportBatch, PurchaseImportRow, Stock
from .serializers import ProductSerializer, PurchaseImportBatchSerializer


def base_product_queryset():
    return Product.objects.select_related("stock").only(
        "id",
        "name",
        "sku",
        "batch",
        "unit",
        "unit_price",
        "is_active",
        "stock__quantity",
    )

HEADER_ALIASES = {
    "name": {"name", "product_name", "product", "productname", "medicine_name", "item_name"},
    "sku": {"sku", "product_sku", "product_code", "code"},
    "batch": {"batch", "batch_no", "batch_number"},
    "unit": {"unit", "uom", "measurement_unit"},
    "unit_price": {"unit_price", "price", "mrp", "rate", "purchase_price"},
    "purchase_quantity": {
        "purchase_quantity",
        "purchase_qty",
        "received_quantity",
        "received_qty",
        "quantity",
        "qty",
        "stock_quantity",
        "stock_qty",
    },
    "is_active": {"is_active", "active", "status"},
}


def normalize_header(value):
    normalized = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower())
    return normalized.strip("_")


def map_headers(headers):
    mapped = {}
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                mapped[field] = index
                break
    return mapped


def parse_decimal(value, field_name, default=None):
    if value in (None, ""):
        return default
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field_name} must be a valid number.")


def parse_integer(value, field_name):
    if value in (None, ""):
        raise ValueError(f"{field_name} is required.")
    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field_name} must be a valid integer.")
    if decimal_value != decimal_value.to_integral_value():
        raise ValueError(f"{field_name} must be a whole number.")
    integer_value = int(decimal_value)
    if integer_value <= 0:
        raise ValueError(f"{field_name} must be greater than zero.")
    return integer_value


def parse_boolean(value, default=True):
    if value in (None, ""):
        return default
    normalized = str(value).strip().lower()
    if normalized in {"true", "1", "yes", "y", "active"}:
        return True
    if normalized in {"false", "0", "no", "n", "inactive"}:
        return False
    raise ValueError("is_active must be true/false, yes/no, or active/inactive.")


def get_cell(row, header_map, field_name):
    index = header_map.get(field_name)
    if index is None or index >= len(row):
        return ""
    return row[index]


def parse_import_file(upload):
    filename = (upload.name or "").lower()
    if filename.endswith(".csv"):
        content = upload.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
    elif filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Excel import requires openpyxl to be installed on the server.") from exc

        workbook = load_workbook(upload, data_only=True)
        worksheet = workbook.active
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    else:
        raise ValueError("Unsupported file type. Please upload a .xlsx or .csv file.")

    if not rows:
        raise ValueError("The uploaded file is empty.")

    headers = [str(value or "").strip() for value in rows[0]]
    header_map = map_headers(headers)
    if "purchase_quantity" not in header_map:
        raise ValueError("The file must include a purchase quantity column.")
    if "sku" not in header_map and "name" not in header_map:
        raise ValueError("The file must include at least a SKU or product name column.")

    return rows[1:], header_map


class ProductListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        search = request.query_params.get("search", "").strip()
        status_filter = request.query_params.get("status", "").strip().lower()

        products = base_product_queryset().order_by("name")
        if search:
            products = products.filter(
                Q(name__icontains=search)
                | Q(sku__icontains=search)
                | Q(batch__icontains=search)
                | Q(unit__icontains=search)
            )
        if status_filter == "active":
            products = products.filter(is_active=True)
        elif status_filter == "inactive":
            products = products.filter(is_active=False)

        serializer = ProductSerializer(products, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = ProductSerializer(data=request.data)
        if serializer.is_valid():
            product = serializer.save()
            return Response(ProductSerializer(product).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProductDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        return Product.objects.get(pk=pk)

    def get(self, request, pk):
        try:
            product = base_product_queryset().get(pk=pk)
            return Response(ProductSerializer(product).data)
        except Product.DoesNotExist:
            return Response({"detail": "Product not found."}, status=status.HTTP_404_NOT_FOUND)

    def patch(self, request, pk):
        try:
            product = self.get_object(pk)
            serializer = ProductSerializer(product, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Product.DoesNotExist:
            return Response({"detail": "Product not found."}, status=status.HTTP_404_NOT_FOUND)

    def put(self, request, pk):
        try:
            product = self.get_object(pk)
            serializer = ProductSerializer(product, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Product.DoesNotExist:
            return Response({"detail": "Product not found."}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, pk):
        try:
            product = self.get_object(pk)
            product.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Product.DoesNotExist:
            return Response({"detail": "Product not found."}, status=status.HTTP_404_NOT_FOUND)


class ProductPurchaseImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "Please choose an Excel or CSV file to import."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows, header_map = parse_import_file(upload)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        summary = {
            "total_rows": 0,
            "processed_rows": 0,
            "created_products": 0,
            "updated_products": 0,
            "failed_rows": 0,
            "total_quantity_added": 0,
        }
        import_batch = PurchaseImportBatch.objects.create(
            uploaded_by=request.user if request.user.is_authenticated else None,
            file_name=upload.name,
        )

        for offset, row in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in row):
                continue

            summary["total_rows"] += 1

            try:
                name = str(get_cell(row, header_map, "name") or "").strip()
                sku = str(get_cell(row, header_map, "sku") or "").strip()
                batch_code = str(get_cell(row, header_map, "batch") or "").strip()
                unit = str(get_cell(row, header_map, "unit") or "Box").strip() or "Box"
                purchase_quantity = parse_integer(get_cell(row, header_map, "purchase_quantity"), "purchase_quantity")
                unit_price = parse_decimal(get_cell(row, header_map, "unit_price"), "unit_price", default=Decimal("0.00"))
                is_active = parse_boolean(get_cell(row, header_map, "is_active"), default=True)

                product = None
                if sku:
                    product = base_product_queryset().filter(sku__iexact=sku).first()
                if product is None and name:
                    product = base_product_queryset().filter(name__iexact=name).first()

                with transaction.atomic():
                    if product:
                        stock, _ = Stock.objects.get_or_create(product=product)
                        stock_before = stock.quantity

                        if name:
                            product.name = name
                        if sku and product.sku != sku:
                            duplicate = Product.objects.filter(sku__iexact=sku).exclude(pk=product.pk).exists()
                            if duplicate:
                                raise ValueError(f'SKU "{sku}" already belongs to another product.')
                            product.sku = sku
                        if batch_code:
                            product.batch = batch_code
                        if unit:
                            product.unit = unit
                        if unit_price is not None:
                            product.unit_price = unit_price
                        product.is_active = is_active
                        product.save()

                        stock.quantity = stock_before + purchase_quantity
                        stock.save(update_fields=["quantity", "updated_at"])

                        summary["processed_rows"] += 1
                        summary["updated_products"] += 1
                        summary["total_quantity_added"] += purchase_quantity
                        PurchaseImportRow.objects.create(
                            batch=import_batch,
                            product=product,
                            row_number=offset,
                            action="updated",
                            product_name=product.name,
                            sku=product.sku,
                            quantity_added=purchase_quantity,
                            stock_before=stock_before,
                            stock_after=stock.quantity,
                            message="Existing stock increased successfully.",
                        )
                    else:
                        if not name:
                            raise ValueError("Product name is required for a new product row.")
                        if not sku:
                            raise ValueError("SKU is required for a new product row.")

                        product = Product.objects.create(
                            name=name,
                            sku=sku,
                            batch=batch_code,
                            unit=unit,
                            unit_price=unit_price if unit_price is not None else Decimal("0.00"),
                            is_active=is_active,
                        )
                        stock = Stock.objects.create(product=product, quantity=purchase_quantity)

                        summary["processed_rows"] += 1
                        summary["created_products"] += 1
                        summary["total_quantity_added"] += purchase_quantity
                        PurchaseImportRow.objects.create(
                            batch=import_batch,
                            product=product,
                            row_number=offset,
                            action="created",
                            product_name=product.name,
                            sku=product.sku,
                            quantity_added=purchase_quantity,
                            stock_before=0,
                            stock_after=stock.quantity,
                            message="New product created and added to stock.",
                        )
            except Exception as exc:
                summary["failed_rows"] += 1
                PurchaseImportRow.objects.create(
                    batch=import_batch,
                    row_number=offset,
                    action="failed",
                    product_name=str(get_cell(row, header_map, "name") or "").strip(),
                    sku=str(get_cell(row, header_map, "sku") or "").strip(),
                    quantity_added=0,
                    stock_before=None,
                    stock_after=None,
                    message=str(exc),
                )

        import_batch.total_rows = summary["total_rows"]
        import_batch.processed_rows = summary["processed_rows"]
        import_batch.created_products = summary["created_products"]
        import_batch.updated_products = summary["updated_products"]
        import_batch.failed_rows = summary["failed_rows"]
        import_batch.total_quantity_added = summary["total_quantity_added"]
        import_batch.save(
            update_fields=[
                "total_rows",
                "processed_rows",
                "created_products",
                "updated_products",
                "failed_rows",
                "total_quantity_added",
            ]
        )

        serializer = PurchaseImportBatchSerializer(import_batch)
        return Response(serializer.data)


class ProductPurchaseImportHistoryListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        limit = request.query_params.get("limit")
        batches = PurchaseImportBatch.objects.select_related("uploaded_by").prefetch_related(
            Prefetch(
                "rows",
                queryset=PurchaseImportRow.objects.select_related("product").only(
                    "id",
                    "batch_id",
                    "product_id",
                    "row_number",
                    "action",
                    "product_name",
                    "sku",
                    "quantity_added",
                    "stock_before",
                    "stock_after",
                    "message",
                ),
            )
        )
        if limit:
            try:
                batches = batches[: max(1, min(int(limit), 100))]
            except ValueError:
                pass
        serializer = PurchaseImportBatchSerializer(batches, many=True)
        return Response(serializer.data)


class ProductPurchaseImportHistoryDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            batch = PurchaseImportBatch.objects.select_related("uploaded_by").prefetch_related(
                Prefetch(
                    "rows",
                    queryset=PurchaseImportRow.objects.select_related("product").only(
                        "id",
                        "batch_id",
                        "product_id",
                        "row_number",
                        "action",
                        "product_name",
                        "sku",
                        "quantity_added",
                        "stock_before",
                        "stock_after",
                        "message",
                    ),
                )
            ).get(pk=pk)
        except PurchaseImportBatch.DoesNotExist:
            return Response({"detail": "Import history record not found."}, status=status.HTTP_404_NOT_FOUND)
        serializer = PurchaseImportBatchSerializer(batch)
        return Response(serializer.data)
