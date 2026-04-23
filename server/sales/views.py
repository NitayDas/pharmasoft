import csv
import io
import re
from datetime import timedelta
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import F, Prefetch
from django.db.models import Count, DecimalField, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.utils import timezone
from rest_framework import status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import PaymentTransaction, Product, Sale, Customer, Stock, PurchaseImportBatch, PurchaseImportRow
from .serializers import (
    CustomerSerializer,
    PaymentTransactionCreateSerializer,
    PaymentTransactionSerializer,
    ProductSerializer,
    PurchaseImportBatchSerializer,
    SaleCreateSerializer,
    SaleInvoiceUpdateSerializer,
    SaleSerializer,
)


def base_product_queryset():
    return Product.objects.select_related('stock').only(
        'id',
        'name',
        'sku',
        'batch',
        'unit',
        'unit_price',
        'is_active',
        'stock__quantity',
    )


def base_sale_queryset():
    sale_item_queryset = Sale._meta.get_field('items').related_model.objects.select_related('product').only(
        'id',
        'sale_id',
        'product_id',
        'item_name',
        'batch',
        'unit',
        'qty',
        'unit_price',
        'discount_percent',
        'gross_amount',
        'discount_amount',
        'net_amount',
        'product__id',
        'product__name',
        'product__sku',
    )
    txn_queryset = PaymentTransaction.objects.select_related('received_by').order_by('-created_at')
    return Sale.objects.select_related('served_by', 'customer').prefetch_related(
        Prefetch('items', queryset=sale_item_queryset),
        Prefetch('transactions', queryset=txn_queryset),
    )


HEADER_ALIASES = {
    'name': {'name', 'product_name', 'product', 'productname', 'medicine_name', 'item_name'},
    'sku': {'sku', 'product_sku', 'product_code', 'code'},
    'batch': {'batch', 'batch_no', 'batch_number'},
    'unit': {'unit', 'uom', 'measurement_unit'},
    'unit_price': {'unit_price', 'price', 'mrp', 'rate', 'purchase_price'},
    'purchase_quantity': {
        'purchase_quantity',
        'purchase_qty',
        'received_quantity',
        'received_qty',
        'quantity',
        'qty',
        'stock_quantity',
        'stock_qty',
    },
    'is_active': {'is_active', 'active', 'status'},
}


def normalize_header(value):
    normalized = re.sub(r'[^a-z0-9]+', '_', str(value or '').strip().lower())
    return normalized.strip('_')


def map_headers(headers):
    mapped = {}
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                mapped[field] = index
                break
    return mapped


def parse_decimal(value, field_name, default=None):
    if value in (None, ''):
        return default
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f'{field_name} must be a valid number.')


def parse_integer(value, field_name):
    if value in (None, ''):
        raise ValueError(f'{field_name} is required.')
    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f'{field_name} must be a valid integer.')
    if decimal_value != decimal_value.to_integral_value():
        raise ValueError(f'{field_name} must be a whole number.')
    integer_value = int(decimal_value)
    if integer_value <= 0:
        raise ValueError(f'{field_name} must be greater than zero.')
    return integer_value


def parse_boolean(value, default=True):
    if value in (None, ''):
        return default
    normalized = str(value).strip().lower()
    if normalized in {'true', '1', 'yes', 'y', 'active'}:
        return True
    if normalized in {'false', '0', 'no', 'n', 'inactive'}:
        return False
    raise ValueError('is_active must be true/false, yes/no, or active/inactive.')


def get_cell(row, header_map, field_name):
    index = header_map.get(field_name)
    if index is None or index >= len(row):
        return ''
    return row[index]


def parse_import_file(upload):
    filename = (upload.name or '').lower()
    if filename.endswith('.csv'):
        content = upload.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
    elif filename.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError('Excel import requires openpyxl to be installed on the server.') from exc

        workbook = load_workbook(upload, data_only=True)
        worksheet = workbook.active
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    else:
        raise ValueError('Unsupported file type. Please upload a .xlsx or .csv file.')

    if not rows:
        raise ValueError('The uploaded file is empty.')

    headers = [str(value or '').strip() for value in rows[0]]
    header_map = map_headers(headers)
    if 'purchase_quantity' not in header_map:
        raise ValueError('The file must include a purchase quantity column.')
    if 'sku' not in header_map and 'name' not in header_map:
        raise ValueError('The file must include at least a SKU or product name column.')

    return rows[1:], header_map


class ProductListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        search = request.query_params.get('search', '').strip()
        status_filter = request.query_params.get('status', '').strip().lower()

        products = base_product_queryset().order_by('name')
        if search:
            products = products.filter(
                Q(name__icontains=search) |
                Q(sku__icontains=search) |
                Q(batch__icontains=search) |
                Q(unit__icontains=search)
            )
        if status_filter == 'active':
            products = products.filter(is_active=True)
        elif status_filter == 'inactive':
            products = products.filter(is_active=False)

        serializer = ProductSerializer(products, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = ProductSerializer(data=request.data)
        if serializer.is_valid():
            product = serializer.save()
            return Response(ProductSerializer(product).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProductDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        return Product.objects.get(pk=pk)

    def get(self, request, pk):
        try:
            product = base_product_queryset().get(pk=pk)
            return Response(ProductSerializer(product).data)
        except Product.DoesNotExist:
            return Response({'detail': 'Product not found.'}, status=status.HTTP_404_NOT_FOUND)


class ProductPurchaseImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'detail': 'Please choose an Excel or CSV file to import.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows, header_map = parse_import_file(upload)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        summary = {
            'total_rows': 0,
            'processed_rows': 0,
            'created_products': 0,
            'updated_products': 0,
            'failed_rows': 0,
            'total_quantity_added': 0,
        }
        import_batch = PurchaseImportBatch.objects.create(
            uploaded_by=request.user if request.user.is_authenticated else None,
            file_name=upload.name,
        )

        for offset, row in enumerate(rows, start=2):
            if not any(value not in (None, '') for value in row):
                continue

            summary['total_rows'] += 1

            try:
                name = str(get_cell(row, header_map, 'name') or '').strip()
                sku = str(get_cell(row, header_map, 'sku') or '').strip()
                batch_code = str(get_cell(row, header_map, 'batch') or '').strip()
                unit = str(get_cell(row, header_map, 'unit') or 'Box').strip() or 'Box'
                purchase_quantity = parse_integer(
                    get_cell(row, header_map, 'purchase_quantity'),
                    'purchase_quantity',
                )
                unit_price = parse_decimal(
                    get_cell(row, header_map, 'unit_price'),
                    'unit_price',
                    default=Decimal('0.00'),
                )
                is_active = parse_boolean(get_cell(row, header_map, 'is_active'), default=True)

                product = None
                if sku:
                    product = base_product_queryset().filter(sku__iexact=sku).first()
                if product is None and name:
                    product = base_product_queryset().filter(name__iexact=name).first()

                with transaction.atomic():
                    if product:
                        stock, _ = Stock.objects.get_or_create(product=product)
                        stock_before = stock.quantity

                        if name:
                            product.name = name
                        if sku and product.sku != sku:
                            duplicate = Product.objects.filter(sku__iexact=sku).exclude(pk=product.pk).exists()
                            if duplicate:
                                raise ValueError(f'SKU "{sku}" already belongs to another product.')
                            product.sku = sku
                        if batch_code:
                            product.batch = batch_code
                        if unit:
                            product.unit = unit
                        if unit_price is not None:
                            product.unit_price = unit_price
                        product.is_active = is_active
                        product.save()

                        stock.quantity = stock_before + purchase_quantity
                        stock.save(update_fields=['quantity', 'updated_at'])

                        summary['processed_rows'] += 1
                        summary['updated_products'] += 1
                        summary['total_quantity_added'] += purchase_quantity
                        PurchaseImportRow.objects.create(
                            batch=import_batch,
                            product=product,
                            row_number=offset,
                            action='updated',
                            product_name=product.name,
                            sku=product.sku,
                            quantity_added=purchase_quantity,
                            stock_before=stock_before,
                            stock_after=stock.quantity,
                            message='Existing stock increased successfully.',
                        )
                    else:
                        if not name:
                            raise ValueError('Product name is required for a new product row.')
                        if not sku:
                            raise ValueError('SKU is required for a new product row.')

                        product = Product.objects.create(
                            name=name,
                            sku=sku,
                            batch=batch_code,
                            unit=unit,
                            unit_price=unit_price if unit_price is not None else Decimal('0.00'),
                            is_active=is_active,
                        )
                        stock = Stock.objects.create(product=product, quantity=purchase_quantity)

                        summary['processed_rows'] += 1
                        summary['created_products'] += 1
                        summary['total_quantity_added'] += purchase_quantity
                        PurchaseImportRow.objects.create(
                            batch=import_batch,
                            product=product,
                            row_number=offset,
                            action='created',
                            product_name=product.name,
                            sku=product.sku,
                            quantity_added=purchase_quantity,
                            stock_before=0,
                            stock_after=stock.quantity,
                            message='New product created and added to stock.',
                        )
            except Exception as exc:
                summary['failed_rows'] += 1
                PurchaseImportRow.objects.create(
                    batch=import_batch,
                    row_number=offset,
                    action='failed',
                    product_name=str(get_cell(row, header_map, 'name') or '').strip(),
                    sku=str(get_cell(row, header_map, 'sku') or '').strip(),
                    quantity_added=0,
                    stock_before=None,
                    stock_after=None,
                    message=str(exc),
                )

        import_batch.total_rows = summary['total_rows']
        import_batch.processed_rows = summary['processed_rows']
        import_batch.created_products = summary['created_products']
        import_batch.updated_products = summary['updated_products']
        import_batch.failed_rows = summary['failed_rows']
        import_batch.total_quantity_added = summary['total_quantity_added']
        import_batch.save(
            update_fields=[
                'total_rows',
                'processed_rows',
                'created_products',
                'updated_products',
                'failed_rows',
                'total_quantity_added',
            ]
        )

        serializer = PurchaseImportBatchSerializer(import_batch)
        return Response(serializer.data)


class ProductPurchaseImportHistoryListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        limit = request.query_params.get('limit')
        batches = PurchaseImportBatch.objects.select_related('uploaded_by').prefetch_related(
            Prefetch(
                'rows',
                queryset=PurchaseImportRow.objects.select_related('product').only(
                    'id',
                    'batch_id',
                    'product_id',
                    'row_number',
                    'action',
                    'product_name',
                    'sku',
                    'quantity_added',
                    'stock_before',
                    'stock_after',
                    'message',
                ),
            )
        )
        if limit:
            try:
                batches = batches[: max(1, min(int(limit), 100))]
            except ValueError:
                pass
        serializer = PurchaseImportBatchSerializer(batches, many=True)
        return Response(serializer.data)


class ProductPurchaseImportHistoryDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            batch = PurchaseImportBatch.objects.select_related('uploaded_by').prefetch_related(
                Prefetch(
                    'rows',
                    queryset=PurchaseImportRow.objects.select_related('product').only(
                        'id',
                        'batch_id',
                        'product_id',
                        'row_number',
                        'action',
                        'product_name',
                        'sku',
                        'quantity_added',
                        'stock_before',
                        'stock_after',
                        'message',
                    ),
                )
            ).get(pk=pk)
        except PurchaseImportBatch.DoesNotExist:
            return Response({'detail': 'Import history record not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = PurchaseImportBatchSerializer(batch)
        return Response(serializer.data)

    def patch(self, request, pk):
        try:
            product = self.get_object(pk)
            serializer = ProductSerializer(product, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Product.DoesNotExist:
            return Response({'detail': 'Product not found.'}, status=status.HTTP_404_NOT_FOUND)

    def put(self, request, pk):
        try:
            product = self.get_object(pk)
            serializer = ProductSerializer(product, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Product.DoesNotExist:
            return Response({'detail': 'Product not found.'}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, pk):
        try:
            product = self.get_object(pk)
            product.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Product.DoesNotExist:
            return Response({'detail': 'Product not found.'}, status=status.HTTP_404_NOT_FOUND)


class SaleCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        sales = base_sale_queryset().order_by('-created_at')
        serializer = SaleSerializer(sales, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = SaleCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        sale = serializer.save()
        return Response(SaleSerializer(sale).data, status=status.HTTP_201_CREATED)


class SaleDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        return base_sale_queryset().get(pk=pk)

    def get(self, request, pk):
        try:
            sale = self.get_object(pk)
        except Sale.DoesNotExist:
            return Response({'detail': 'Invoice not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(SaleSerializer(sale).data)

    def patch(self, request, pk):
        try:
            sale = self.get_object(pk)
        except Sale.DoesNotExist:
            return Response({'detail': 'Invoice not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = SaleInvoiceUpdateSerializer(sale, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        updated_sale = serializer.save()
        refreshed_sale = base_sale_queryset().get(pk=updated_sale.pk)
        return Response(SaleSerializer(refreshed_sale).data)


class LatestSaleView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        sale = base_sale_queryset().order_by('-created_at').first()
        if not sale:
            return Response({'sale': None})
        return Response({'sale': SaleSerializer(sale).data})


class DashboardSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = timezone.localdate()
        expiry_threshold = today + timedelta(days=30)

        sales_today = Sale.objects.filter(sale_date=today)
        summary = sales_today.aggregate(
            total_sales=Coalesce(
                Sum('grand_total'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            total_discount=Coalesce(
                Sum('discount_amount'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
            sale_count=Count('id'),
        )

        today_collected = PaymentTransaction.objects.filter(payment_date=today).aggregate(
            total=Coalesce(
                Sum('amount'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )['total']

        total_outstanding_due = Sale.objects.aggregate(
            total=Coalesce(
                Sum('due_amount'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )['total']

        low_stock_count = Stock.objects.filter(
            quantity__lte=F('product__reorder_level')
        ).count()

        expiring_soon_count = Product.objects.filter(
            is_active=True,
            expiry_date__isnull=False,
            expiry_date__lte=expiry_threshold,
            expiry_date__gte=today,
        ).count()

        expired_count = Product.objects.filter(
            is_active=True,
            expiry_date__isnull=False,
            expiry_date__lt=today,
        ).count()

        latest_sale = base_sale_queryset().order_by('-created_at').first()

        return Response({
            'today': str(today),
            'total_sales': summary['total_sales'],
            'total_discount': summary['total_discount'],
            'sale_count': summary['sale_count'],
            'today_collected': today_collected,
            'total_outstanding_due': total_outstanding_due,
            'low_stock_count': low_stock_count,
            'expiring_soon_count': expiring_soon_count,
            'expired_count': expired_count,
            'latest_sale': SaleSerializer(latest_sale).data if latest_sale else None,
        })


class InvoicePaymentListCreateView(APIView):
    """List or record payments for a specific invoice."""
    permission_classes = [IsAuthenticated]

    def _get_sale(self, pk):
        return Sale.objects.get(pk=pk)

    def get(self, request, pk):
        try:
            sale = self._get_sale(pk)
        except Sale.DoesNotExist:
            return Response({'detail': 'Invoice not found.'}, status=status.HTTP_404_NOT_FOUND)
        txns = PaymentTransaction.objects.filter(sale=sale).select_related('received_by').order_by('-created_at')
        return Response(PaymentTransactionSerializer(txns, many=True).data)

    @transaction.atomic
    def post(self, request, pk):
        try:
            sale = Sale.objects.select_for_update().get(pk=pk)
        except Sale.DoesNotExist:
            return Response({'detail': 'Invoice not found.'}, status=status.HTTP_404_NOT_FOUND)
        if sale.due_amount <= 0:
            return Response({'detail': 'Invoice is already fully paid.'}, status=status.HTTP_400_BAD_REQUEST)
        serializer = PaymentTransactionCreateSerializer(
            data=request.data,
            context={'sale': sale, 'request': request},
        )
        serializer.is_valid(raise_exception=True)
        txn = serializer.save()
        return Response(PaymentTransactionSerializer(txn).data, status=status.HTTP_201_CREATED)


class SalesPaymentListView(APIView):
    """List all sales with payment status — used by the Sales/Payments page."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_from = request.query_params.get('date_from', '').strip()
        date_to = request.query_params.get('date_to', '').strip()
        payment_method = request.query_params.get('payment_method', '').strip()
        payment_status_filter = request.query_params.get('payment_status', '').strip().lower()
        search = request.query_params.get('search', '').strip()

        sales = base_sale_queryset()

        if date_from:
            sales = sales.filter(sale_date__gte=date_from)
        if date_to:
            sales = sales.filter(sale_date__lte=date_to)
        if payment_method:
            sales = sales.filter(payment_method=payment_method)
        if payment_status_filter == 'paid':
            sales = sales.filter(due_amount__lte=0)
        elif payment_status_filter == 'partial':
            sales = sales.filter(paid_amount__gt=0, due_amount__gt=0)
        elif payment_status_filter == 'unpaid':
            sales = sales.filter(paid_amount=0, due_amount__gt=0)
        if search:
            sales = sales.filter(
                Q(sale_no__icontains=search) |
                Q(customer_name__icontains=search) |
                Q(contact_number__icontains=search)
            )

        sales = sales.order_by('-sale_date', '-created_at')
        return Response(SaleSerializer(sales, many=True).data)


class CustomerLedgerView(APIView):
    """Full purchase + payment ledger for a customer."""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            customer = Customer.objects.get(pk=pk)
        except Customer.DoesNotExist:
            return Response({'detail': 'Customer not found.'}, status=status.HTTP_404_NOT_FOUND)

        sales = base_sale_queryset().filter(customer=customer).order_by('-sale_date', '-created_at')
        sales_data = SaleSerializer(sales, many=True).data

        total_paid = PaymentTransaction.objects.filter(sale__customer=customer).aggregate(
            total=Coalesce(
                Sum('amount'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )['total']

        return Response({
            'customer': CustomerSerializer(customer).data,
            'invoices': sales_data,
            'summary': {
                'total_invoices': sales.count(),
                'total_purchase': str(customer.total_purchase_amount),
                'total_paid': str(total_paid),
                'total_due': str(customer.total_due_amount),
            },
        })


class CustomerCreateView(APIView):
    """Create a new customer."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = CustomerSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CustomerGetByPhoneView(APIView):
    """Get customer by phone number."""
    permission_classes = [IsAuthenticated]

    def get(self, request, phone):
        try:
            customer = Customer.objects.get(phone=phone)
            serializer = CustomerSerializer(customer)
            return Response(serializer.data)
        except Customer.DoesNotExist:
            return Response({'detail': 'Customer not found'}, status=status.HTTP_404_NOT_FOUND)


class CustomerListView(APIView):
    """List all customers."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        search = request.query_params.get('search', '').strip()
        customers = Customer.objects.only(
            'id',
            'name',
            'phone',
            'email',
            'address',
            'total_purchase_amount',
            'total_due_amount',
            'medicine_history',
            'created_at',
            'updated_at',
        ).order_by('name')
        if search:
            customers = customers.filter(
                Q(name__icontains=search) |
                Q(phone__icontains=search) |
                Q(email__icontains=search)
            )
        serializer = CustomerSerializer(customers, many=True)
        return Response(serializer.data)


class CustomerDetailView(APIView):
    """Retrieve, update, or delete a customer."""
    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        return Customer.objects.get(pk=pk)

    def get(self, request, pk):
        try:
            customer = self.get_object(pk)
            serializer = CustomerSerializer(customer)
            return Response(serializer.data)
        except Customer.DoesNotExist:
            return Response({'detail': 'Customer not found'}, status=status.HTTP_404_NOT_FOUND)

    def patch(self, request, pk):
        try:
            customer = self.get_object(pk)
            serializer = CustomerSerializer(customer, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Customer.DoesNotExist:
            return Response({'detail': 'Customer not found'}, status=status.HTTP_404_NOT_FOUND)

    def put(self, request, pk):
        try:
            customer = self.get_object(pk)
            serializer = CustomerSerializer(customer, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Customer.DoesNotExist:
            return Response({'detail': 'Customer not found'}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, pk):
        try:
            customer = self.get_object(pk)
            customer.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Customer.DoesNotExist:
            return Response({'detail': 'Customer not found'}, status=status.HTTP_404_NOT_FOUND)
