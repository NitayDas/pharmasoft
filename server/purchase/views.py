import csv
import io
import re
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Prefetch, Q
from rest_framework import status
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from sales.models import Product, Stock

from .models import PurchaseEntry, PurchaseEntryItem
from .serializers import ManualPurchaseCreateSerializer, PurchaseEntrySerializer


def base_product_queryset():
    return Product.objects.select_related("stock").only(
        "id",
        "name",
        "sku",
        "batch",
        "unit",
        "unit_price",
        "is_active",
        "stock__quantity",
    )


def base_purchase_queryset():
    item_queryset = PurchaseEntryItem.objects.select_related("product").only(
        "id",
        "purchase_entry_id",
        "product_id",
        "row_number",
        "action",
        "product_name",
        "sku",
        "batch",
        "unit",
        "unit_price",
        "quantity_added",
        "stock_before",
        "stock_after",
        "message",
        "product__id",
        "product__name",
        "product__sku",
    )
    return PurchaseEntry.objects.select_related("created_by").prefetch_related(
        Prefetch("items", queryset=item_queryset)
    )


HEADER_ALIASES = {
    "name": {"name", "product_name", "product", "productname", "medicine_name", "item_name"},
    "sku": {"sku", "product_sku", "product_code", "code"},
    "batch": {"batch", "batch_no", "batch_number"},
    "unit": {"unit", "uom", "measurement_unit"},
    "unit_price": {"unit_price", "price", "mrp", "rate", "purchase_price"},
    "purchase_quantity": {
        "purchase_quantity",
        "purchase_qty",
        "received_quantity",
        "received_qty",
        "quantity",
        "qty",
        "stock_quantity",
        "stock_qty",
    },
    "is_active": {"is_active", "active", "status"},
}


def normalize_header(value):
    normalized = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower())
    return normalized.strip("_")


def map_headers(headers):
    mapped = {}
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        for field, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                mapped[field] = index
                break
    return mapped


def parse_decimal(value, field_name, default=None):
    if value in (None, ""):
        return default
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field_name} must be a valid number.")


def parse_integer(value, field_name):
    if value in (None, ""):
        raise ValueError(f"{field_name} is required.")
    try:
        decimal_value = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f"{field_name} must be a valid integer.")
    if decimal_value != decimal_value.to_integral_value():
        raise ValueError(f"{field_name} must be a whole number.")
    integer_value = int(decimal_value)
    if integer_value <= 0:
        raise ValueError(f"{field_name} must be greater than zero.")
    return integer_value


def parse_boolean(value, default=True):
    if value in (None, ""):
        return default
    normalized = str(value).strip().lower()
    if normalized in {"true", "1", "yes", "y", "active"}:
        return True
    if normalized in {"false", "0", "no", "n", "inactive"}:
        return False
    raise ValueError("is_active must be true/false, yes/no, or active/inactive.")


def get_cell(row, header_map, field_name):
    index = header_map.get(field_name)
    if index is None or index >= len(row):
        return ""
    return row[index]


def parse_import_file(upload):
    filename = (upload.name or "").lower()
    if filename.endswith(".csv"):
        content = upload.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
    elif filename.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Excel import requires openpyxl to be installed on the server.") from exc

        workbook = load_workbook(upload, data_only=True)
        worksheet = workbook.active
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    else:
        raise ValueError("Unsupported file type. Please upload a .xlsx or .csv file.")

    if not rows:
        raise ValueError("The uploaded file is empty.")

    headers = [str(value or "").strip() for value in rows[0]]
    header_map = map_headers(headers)
    if "purchase_quantity" not in header_map:
        raise ValueError("The file must include a purchase quantity column.")
    if "sku" not in header_map and "name" not in header_map:
        raise ValueError("The file must include at least a SKU or product name column.")

    return rows[1:], header_map


def serialize_purchase_entry(entry):
    return PurchaseEntrySerializer(base_purchase_queryset().get(pk=entry.pk)).data


class PurchaseEntryListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        search = request.query_params.get("search", "").strip()
        source = request.query_params.get("source", "").strip().lower()

        purchases = base_purchase_queryset()
        if source in {"excel", "manual"}:
            purchases = purchases.filter(source=source)
        if search:
            purchases = purchases.filter(
                Q(entry_no__icontains=search)
                | Q(file_name__icontains=search)
                | Q(notes__icontains=search)
                | Q(created_by__username__icontains=search)
                | Q(items__product_name__icontains=search)
                | Q(items__sku__icontains=search)
            ).distinct()

        return Response(PurchaseEntrySerializer(purchases, many=True).data)


class PurchaseEntryDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            purchase_entry = base_purchase_queryset().get(pk=pk)
        except PurchaseEntry.DoesNotExist:
            return Response({"detail": "Purchase entry not found."}, status=status.HTTP_404_NOT_FOUND)

        return Response(PurchaseEntrySerializer(purchase_entry).data)


class PurchaseManualEntryView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        serializer = ManualPurchaseCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data

        name = validated_data["name"].strip()
        sku = validated_data["sku"].strip()
        batch_code = validated_data.get("batch", "").strip()
        unit = (validated_data.get("unit", "") or "Box").strip() or "Box"
        unit_price = validated_data.get("unit_price", Decimal("0.00"))
        purchase_quantity = validated_data["purchase_quantity"]
        is_active = validated_data.get("is_active", True)
        notes = validated_data.get("notes", "").strip()

        product = base_product_queryset().filter(sku__iexact=sku).first()
        if product is None:
            product = base_product_queryset().filter(name__iexact=name).first()

        purchase_entry = PurchaseEntry.objects.create(
            source="manual",
            notes=notes,
            created_by=request.user if request.user.is_authenticated else None,
            total_rows=1,
            processed_rows=1,
            total_quantity_added=purchase_quantity,
        )

        if product:
            stock, _ = Stock.objects.get_or_create(product=product)
            stock_before = stock.quantity

            product.name = name
            if sku and product.sku != sku:
                duplicate = Product.objects.filter(sku__iexact=sku).exclude(pk=product.pk).exists()
                if duplicate:
                    return Response({"detail": f'SKU "{sku}" already belongs to another product.'}, status=status.HTTP_400_BAD_REQUEST)
                product.sku = sku
            product.batch = batch_code
            product.unit = unit
            product.unit_price = unit_price
            product.is_active = is_active
            product.save()

            stock.quantity = stock_before + purchase_quantity
            stock.save(update_fields=["quantity", "updated_at"])

            purchase_entry.updated_products = 1
            purchase_entry.save(update_fields=["updated_products"])

            PurchaseEntryItem.objects.create(
                purchase_entry=purchase_entry,
                product=product,
                row_number=1,
                action="updated",
                product_name=product.name,
                sku=product.sku,
                batch=product.batch,
                unit=product.unit,
                unit_price=product.unit_price,
                quantity_added=purchase_quantity,
                stock_before=stock_before,
                stock_after=stock.quantity,
                message="Manual purchase added to existing stock.",
            )
        else:
            product = Product.objects.create(
                name=name,
                sku=sku,
                batch=batch_code,
                unit=unit,
                unit_price=unit_price,
                is_active=is_active,
            )
            stock = Stock.objects.create(product=product, quantity=purchase_quantity)

            purchase_entry.created_products = 1
            purchase_entry.save(update_fields=["created_products"])

            PurchaseEntryItem.objects.create(
                purchase_entry=purchase_entry,
                product=product,
                row_number=1,
                action="created",
                product_name=product.name,
                sku=product.sku,
                batch=product.batch,
                unit=product.unit,
                unit_price=product.unit_price,
                quantity_added=purchase_quantity,
                stock_before=0,
                stock_after=stock.quantity,
                message="Manual purchase created a new product and stock entry.",
            )

        return Response(serialize_purchase_entry(purchase_entry), status=status.HTTP_201_CREATED)


class PurchaseExcelImportView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    @transaction.atomic
    def post(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "Please choose an Excel or CSV file to import."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows, header_map = parse_import_file(upload)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        purchase_entry = PurchaseEntry.objects.create(
            source="excel",
            file_name=upload.name,
            created_by=request.user if request.user.is_authenticated else None,
        )

        summary = {
            "total_rows": 0,
            "processed_rows": 0,
            "created_products": 0,
            "updated_products": 0,
            "failed_rows": 0,
            "total_quantity_added": 0,
        }

        for offset, row in enumerate(rows, start=2):
            if not any(value not in (None, "") for value in row):
                continue

            summary["total_rows"] += 1

            try:
                name = str(get_cell(row, header_map, "name") or "").strip()
                sku = str(get_cell(row, header_map, "sku") or "").strip()
                batch_code = str(get_cell(row, header_map, "batch") or "").strip()
                unit = str(get_cell(row, header_map, "unit") or "Box").strip() or "Box"
                purchase_quantity = parse_integer(get_cell(row, header_map, "purchase_quantity"), "purchase_quantity")
                unit_price = parse_decimal(get_cell(row, header_map, "unit_price"), "unit_price", default=Decimal("0.00"))
                is_active = parse_boolean(get_cell(row, header_map, "is_active"), default=True)

                product = None
                if sku:
                    product = base_product_queryset().filter(sku__iexact=sku).first()
                if product is None and name:
                    product = base_product_queryset().filter(name__iexact=name).first()

                if product:
                    stock, _ = Stock.objects.get_or_create(product=product)
                    stock_before = stock.quantity

                    if name:
                        product.name = name
                    if sku and product.sku != sku:
                        duplicate = Product.objects.filter(sku__iexact=sku).exclude(pk=product.pk).exists()
                        if duplicate:
                            raise ValueError(f'SKU "{sku}" already belongs to another product.')
                        product.sku = sku
                    if batch_code:
                        product.batch = batch_code
                    if unit:
                        product.unit = unit
                    if unit_price is not None:
                        product.unit_price = unit_price
                    product.is_active = is_active
                    product.save()

                    stock.quantity = stock_before + purchase_quantity
                    stock.save(update_fields=["quantity", "updated_at"])

                    summary["processed_rows"] += 1
                    summary["updated_products"] += 1
                    summary["total_quantity_added"] += purchase_quantity

                    PurchaseEntryItem.objects.create(
                        purchase_entry=purchase_entry,
                        product=product,
                        row_number=offset,
                        action="updated",
                        product_name=product.name,
                        sku=product.sku,
                        batch=product.batch,
                        unit=product.unit,
                        unit_price=product.unit_price,
                        quantity_added=purchase_quantity,
                        stock_before=stock_before,
                        stock_after=stock.quantity,
                        message="Existing stock increased successfully.",
                    )
                else:
                    if not name:
                        raise ValueError("Product name is required for a new product row.")
                    if not sku:
                        raise ValueError("SKU is required for a new product row.")

                    product = Product.objects.create(
                        name=name,
                        sku=sku,
                        batch=batch_code,
                        unit=unit,
                        unit_price=unit_price if unit_price is not None else Decimal("0.00"),
                        is_active=is_active,
                    )
                    stock = Stock.objects.create(product=product, quantity=purchase_quantity)

                    summary["processed_rows"] += 1
                    summary["created_products"] += 1
                    summary["total_quantity_added"] += purchase_quantity

                    PurchaseEntryItem.objects.create(
                        purchase_entry=purchase_entry,
                        product=product,
                        row_number=offset,
                        action="created",
                        product_name=product.name,
                        sku=product.sku,
                        batch=product.batch,
                        unit=product.unit,
                        unit_price=product.unit_price,
                        quantity_added=purchase_quantity,
                        stock_before=0,
                        stock_after=stock.quantity,
                        message="New product created and added to stock.",
                    )
            except Exception as exc:
                summary["failed_rows"] += 1
                PurchaseEntryItem.objects.create(
                    purchase_entry=purchase_entry,
                    row_number=offset,
                    action="failed",
                    product_name=str(get_cell(row, header_map, "name") or "").strip(),
                    sku=str(get_cell(row, header_map, "sku") or "").strip(),
                    batch=str(get_cell(row, header_map, "batch") or "").strip(),
                    unit=str(get_cell(row, header_map, "unit") or "").strip(),
                    unit_price=Decimal("0.00"),
                    quantity_added=0,
                    stock_before=None,
                    stock_after=None,
                    message=str(exc),
                )

        purchase_entry.total_rows = summary["total_rows"]
        purchase_entry.processed_rows = summary["processed_rows"]
        purchase_entry.created_products = summary["created_products"]
        purchase_entry.updated_products = summary["updated_products"]
        purchase_entry.failed_rows = summary["failed_rows"]
        purchase_entry.total_quantity_added = summary["total_quantity_added"]
        purchase_entry.save(
            update_fields=[
                "total_rows",
                "processed_rows",
                "created_products",
                "updated_products",
                "failed_rows",
                "total_quantity_added",
            ]
        )

        return Response(serialize_purchase_entry(purchase_entry), status=status.HTTP_201_CREATED)
